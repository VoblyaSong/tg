import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram import Router
import re

TOKEN = '8117709260:AAGscZ7bD-IC5qoQNdTiA9cYbB0ZxvWaZ6A'

bot = Bot(TOKEN)
dp = Dispatcher()
router = Router()
dp.include_router(router)

countries = ["Франция", "Германия", "Грузия", "Италия", "Испания", "Швеция"]

class VotingStates(StatesGroup):
    choosing_country = State()
    sending_scores = State()

@router.message(F.text == "/start")
async def cmd_start(message: types.Message, state: FSMContext):
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
        [KeyboardButton(text=country)] for country in countries
    ] + [[KeyboardButton(text="Отмена")]])
    await state.set_state(VotingStates.choosing_country)
    await message.answer("Жюри какой страны вы представляете?", reply_markup=keyboard)

    @router.message(F.text.lower() == "отмена")
    async def cancel_action(message: types.Message, state: FSMContext):
        await state.clear()
        await message.answer("Действие отменено. Чтобы начать заново, введите /start",
                             reply_markup=types.ReplyKeyboardRemove())

@router.message(VotingStates.choosing_country)
async def handle_country_choice(message: types.Message, state: FSMContext):
    if message.text not in countries:
        await message.answer("Пожалуйста, выберите страну из списка или нажмите 'Отмена'.")
        return
    await state.update_data(country=message.text)
    await state.set_state(VotingStates.sending_scores)
    await message.answer(
        "Прекрасно! Теперь отправляйте свои голоса по системе Евровидения (1-8, 10, 12)\n\n"
        "Формат:\n12 - Страна\n10 - Страна\n...\n\n",
        reply_markup=types.ReplyKeyboardRemove()
    )


@router.message(VotingStates.sending_scores)
async def handle_scores(message: types.Message, state: FSMContext):
    text = message.text.strip()
    pattern = r"(1[0-2]|[1-9])\s*-\s*(.+)"
    matches = re.findall(pattern, text)

    expected_points = {'12', '10', '8', '7', '6', '5', '4', '3', '2', '1'}
    submitted_points = set()
    countries_voted = []

    for point, country in matches:
        submitted_points.add(point)
        countries_voted.append(country.strip().lower())

    missing = sorted(expected_points - submitted_points, reverse=True)

    if missing:
        await message.answer(f"⚠️ Пожалуйста, отправьте все баллы! Отсутствуют: {', '.join(missing)}")
        return

    duplicates = [country for country in countries_voted if countries_voted.count(country) > 1]
    if duplicates:
        unique_duplicates = sorted(set(duplicates))
        await message.answer(f"⚠️ Страны не должны повторяться. Повторы: {', '.join(unique_duplicates)}")
        return

    import openpyxl
    from openpyxl.utils import get_column_letter
    import os

    file_name = "votes.xlsx"
    user_data = await state.get_data()
    jury_country = user_data.get('country', 'Неизвестно')
    user_id = message.from_user.id

    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Votes"
        ws.append(["Жюри", "User ID", "Баллы"])
    else:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

    votes_str = ", ".join([f"{point} - {country.strip()}" for point, country in matches])

    ws.append([jury_country, user_id, votes_str])
    wb.save(file_name)

    await message.answer("Спасибо, баллы приняты.")
    await state.clear()


async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
